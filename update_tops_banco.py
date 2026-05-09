import openpyxl, sys
sys.path.insert(0, r'C:\fleet-cloud')

from app.database import engine_sync
from sqlalchemy import text

# Ler arquivo
wb = openpyxl.load_workbook(r'C:\Users\victor.mosqueira\Downloads\Cab_0805.xlsx')
ws = wb.active
headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows())]

id_idx = headers.index('Nro. Nota')
top_idx = headers.index('Tipo Operação')
vlr_idx = headers.index('Vlr. Nota')

with engine_sync.connect() as conn:
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[id_idx]:
            continue
        external_id = str(row[id_idx])
        order_type = str(row[top_idx]) if row[top_idx] else '1000'
        total_value = float(row[vlr_idx]) if row[vlr_idx] else None

        conn.execute(text(
            "UPDATE orders SET order_type=:ot, total_value=:tv WHERE external_id=:eid"
        ), {"ot": order_type, "tv": total_value, "eid": external_id})
        updated += 1

    conn.commit()
    print(f"OK! {updated} pedidos atualizados!")

    # Verificar resultado
    r = conn.execute(text(
        "SELECT order_type, COUNT(*) as qtd FROM orders GROUP BY order_type"
    ))
    for row in r:
        print(f"  TOP {row[0]}: {row[1]} pedidos")
