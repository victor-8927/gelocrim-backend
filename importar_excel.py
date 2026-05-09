"""
importar_excel.py v2 - Corrigido para dados reais do Sankhya
STATUS_PEDIDO: 'Sim' = liberado
CONFIRMADO: 'Sim'
TEM_ORDEM_CARGA: 0 = sem ordem
PESOBRUT pode ser None, usa PESOLIQ
"""
import sys, os
from uuid import uuid4
from datetime import datetime, timezone

sys.path.insert(0, r'C:\fleet-cloud')
os.chdir(r'C:\fleet-cloud')

import openpyxl
from app.database import engine_sync, init_schema
from sqlalchemy import text

def now_str():
    return datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

def parse_float(val, default=0.0):
    try:
        return float(str(val).replace(',', '.')) if val is not None else default
    except:
        return default

def parse_time(val, default='08:00'):
    if not val:
        return default
    s = str(val).strip()
    if ':' in s:
        return s[:5]
    return default

arquivo = sys.argv[1] if len(sys.argv) > 1 else r'C:\fleet-cloud\Gelocrim_Importacao_Dados.xlsx'

if not os.path.exists(arquivo):
    print(f'Arquivo nao encontrado: {arquivo}')
    sys.exit(1)

print(f'\nImportando: {arquivo}')
print('=' * 60)

wb = openpyxl.load_workbook(arquivo, data_only=True)
print(f'Abas: {wb.sheetnames}')

init_schema()

stats = {
    'clientes': {'ok': 0, 'skip': 0, 'erro': 0},
    'pedidos':  {'ok': 0, 'skip': 0, 'erro': 0},
    'veiculos': {'ok': 0, 'skip': 0, 'erro': 0},
}

with engine_sync.begin() as conn:

    # CLIENTES
    sheet_cli = None
    for name in wb.sheetnames:
        if 'client' in name.lower() or 'tgfpar' in name.lower():
            sheet_cli = wb[name]
            break

    if sheet_cli:
        print(f'\nImportando clientes: {sheet_cli.title}')
        for row in sheet_cli.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            codparc = str(row[0]).strip()
            nome = str(row[1] or '').strip()
            if not nome or nome == 'NOMEPARC *':
                continue
            try:
                ext_id = f'SNK-CLI-{codparc}'
                partes = [str(row[5] or ''), str(row[6] or ''), str(row[8] or '')]
                endereco = ', '.join(p for p in partes if p and p != 'None')
                cidade = str(row[10] or 'Manaus')
                estado = str(row[11] or 'AM')
                if cidade:
                    endereco += f', {cidade}-{estado}'
                lat = parse_float(row[12], -3.1019)
                lng = parse_float(row[13], -60.0250)
                phone = str(row[4] or '')
                ex = conn.execute(text('SELECT id FROM recipients WHERE id=:id'), {'id': ext_id}).fetchone()
                if ex:
                    conn.execute(text('UPDATE recipients SET name=:n,address=:a,lat=:lat,lng=:lng,phone=:p WHERE id=:id'),
                        {'id': ext_id, 'n': nome, 'a': endereco or 'Manaus-AM', 'lat': lat, 'lng': lng, 'p': phone})
                    stats['clientes']['skip'] += 1
                else:
                    conn.execute(text('INSERT INTO recipients (id,name,address,lat,lng,phone,created_at) VALUES (:id,:n,:a,:lat,:lng,:p,:ts)'),
                        {'id': ext_id, 'n': nome, 'a': endereco or 'Manaus-AM', 'lat': lat, 'lng': lng, 'p': phone, 'ts': now_str()})
                    stats['clientes']['ok'] += 1
            except Exception as e:
                stats['clientes']['erro'] += 1

    # PEDIDOS
    sheet_ped = None
    for name in wb.sheetnames:
        if 'pedido' in name.lower() or 'tgfcab' in name.lower():
            sheet_ped = wb[name]
            break

    if sheet_ped:
        print(f'\nImportando pedidos: {sheet_ped.title}')
        ignorados = {'status': 0, 'confirmado': 0, 'ordem_carga': 0, 'sem_cliente': 0, 'duplicado': 0}

        for row in sheet_ped.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            nunota = str(row[0]).strip()
            if not nunota or nunota == 'NUNOTA *':
                continue
            codparc = str(row[2] or '').strip()

            # Regras de negocio
            status_val     = str(row[10] or '').strip().upper()
            confirmado_val = str(row[11] or '').strip().upper()
            ordem_val      = str(row[12] if row[12] is not None else '0').strip().upper()

            # STATUS ok: Sim, S, 1, PENDENTE, LIBERADO
            if status_val not in ('SIM', 'S', '1', 'PENDENTE', 'LIBERADO', 'TRUE'):
                ignorados['status'] += 1
                stats['pedidos']['skip'] += 1
                continue

            # CONFIRMADO ok: Sim, S, 1
            if confirmado_val not in ('SIM', 'S', '1', 'TRUE'):
                ignorados['confirmado'] += 1
                stats['pedidos']['skip'] += 1
                continue

            # SEM ORDEM DE CARGA: 0, NAO, '', None
            if ordem_val in ('1', 'SIM', 'S', 'TRUE'):
                ignorados['ordem_carga'] += 1
                stats['pedidos']['skip'] += 1
                continue

            try:
                ext_id = f'SNK-{nunota}'
                ex = conn.execute(text('SELECT id FROM orders WHERE external_id=:e'), {'e': ext_id}).fetchone()
                if ex:
                    ignorados['duplicado'] += 1
                    stats['pedidos']['skip'] += 1
                    continue

                rec = conn.execute(text('SELECT id,lat,lng FROM recipients WHERE id=:id'),
                    {'id': f'SNK-CLI-{codparc}'}).fetchone()
                if not rec:
                    ignorados['sem_cliente'] += 1
                    stats['pedidos']['skip'] += 1
                    continue

                # Peso: PESOBRUT ou PESOLIQ
                peso = parse_float(row[5], 0) or parse_float(row[6], 0)
                volume = parse_float(row[7], 0)
                tws = parse_time(row[13], '07:30')
                twe = parse_time(row[14], '18:00')
                prio = int(parse_float(row[15], 1)) or 1
                obs = str(row[16] or '')

                data_ent = None
                if row[4]:
                    try:
                        if hasattr(row[4], 'strftime'):
                            data_ent = row[4].strftime('%Y-%m-%d')
                        else:
                            parts = str(row[4]).split('/')
                            if len(parts) == 3:
                                data_ent = f'{parts[2]}-{parts[1]}-{parts[0]}'
                    except:
                        pass

                conn.execute(text('''
                    INSERT INTO orders (id,external_id,source,recipient_id,lat,lng,
                        weight_kg,volume_m3,tw_start,tw_end,priority,status,notes,delivery_date,created_at,updated_at)
                    VALUES (:id,:ext,'sankhya_excel',:rid,:lat,:lng,
                        :kg,:m3,:tws,:twe,:prio,'pending',:notes,:ddate,:ts,:ts)
                '''), {'id': str(uuid4()), 'ext': ext_id, 'rid': rec.id,
                       'lat': rec.lat, 'lng': rec.lng, 'kg': peso, 'm3': volume,
                       'tws': tws, 'twe': twe, 'prio': prio, 'notes': obs,
                       'ddate': data_ent, 'ts': now_str()})
                stats['pedidos']['ok'] += 1
                if stats['pedidos']['ok'] % 20 == 0:
                    print(f'  ... {stats["pedidos"]["ok"]} pedidos importados')

            except Exception as e:
                stats['pedidos']['erro'] += 1
                print(f'  ! Erro {nunota}: {e}')

        print(f'  Ignorados: status={ignorados["status"]} | confirmado={ignorados["confirmado"]} | ordem_carga={ignorados["ordem_carga"]} | sem_cliente={ignorados["sem_cliente"]} | duplicado={ignorados["duplicado"]}')

    # VEICULOS
    sheet_vei = None
    for name in wb.sheetnames:
        if 'vei' in name.lower() or 'frota' in name.lower():
            sheet_vei = wb[name]
            break

    if sheet_vei:
        print(f'\nImportando veiculos: {sheet_vei.title}')
        for row in sheet_vei.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            placa = str(row[0]).strip().upper()
            if not placa or 'PLACA' in placa or 'TOTAL' in placa:
                continue
            try:
                modelo = str(row[1] or placa)
                tipo_str = str(row[2] or '').lower()
                tipo = 'van' if 'van' in tipo_str else 'moto' if 'moto' in tipo_str else 'truck'
                cap_kg = parse_float(row[3], 1000)
                cap_m3 = parse_float(row[4], 8)
                status = 'active' if 'ATIVO' in str(row[5] or 'ATIVO').upper() else 'inactive'
                ex = conn.execute(text('SELECT id FROM vehicles WHERE plate=:p'), {'p': placa}).fetchone()
                if ex:
                    conn.execute(text('UPDATE vehicles SET model=:m,type=:t,capacity_kg=:kg,capacity_m3=:m3,status=:s,updated_at=:ts WHERE plate=:p'),
                        {'m': modelo, 't': tipo, 'kg': cap_kg, 'm3': cap_m3, 's': status, 'p': placa, 'ts': now_str()})
                    stats['veiculos']['skip'] += 1
                else:
                    conn.execute(text('INSERT INTO vehicles (id,plate,model,type,capacity_kg,capacity_m3,status,created_at,updated_at) VALUES (:id,:p,:m,:t,:kg,:m3,:s,:ts,:ts)'),
                        {'id': str(uuid4()), 'p': placa, 'm': modelo, 't': tipo, 'kg': cap_kg, 'm3': cap_m3, 's': status, 'ts': now_str()})
                    stats['veiculos']['ok'] += 1
                    print(f'  + {placa} | {modelo} | {cap_kg}kg')
            except Exception as e:
                stats['veiculos']['erro'] += 1
                print(f'  ! Erro {row[0]}: {e}')

print('\n' + '=' * 60)
print('RESULTADO FINAL:')
print(f"  Clientes : {stats['clientes']['ok']:5d} criados | {stats['clientes']['skip']:5d} ja existiam")
print(f"  Pedidos  : {stats['pedidos']['ok']:5d} criados | {stats['pedidos']['skip']:5d} ignorados")
print(f"  Veiculos : {stats['veiculos']['ok']:5d} criados | {stats['veiculos']['skip']:5d} atualizados")
print('=' * 60)
if stats['pedidos']['ok'] > 0:
    print(f'\nSucesso! Acesse o sistema > Pedidos para ver os dados.')
else:
    print('\nNenhum pedido importado. Verifique os dados na planilha.')
