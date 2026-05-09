import openpyxl
wb = openpyxl.load_workbook('Gelocrim_Importacao_Dados.xlsx', data_only=True)
print('Abas:', wb.sheetnames)
for name in wb.sheetnames:
    if 'ped' in name.lower() or 'tgfcab' in name.lower():
        ws = wb[name]
        print(f'\nAba: {name}')
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True)):
            print(f'  Linha {i+2}: {row}')
        break
